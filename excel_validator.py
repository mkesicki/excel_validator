#!/usr/bin/python -u
# -*- coding: UTF-8 -*-
import argparse
import os.path
import sys
import time
import yaml
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from progress.bar import Bar
from validator import *

def isValid(type, value, coordinate, errors, value2 = None):
    '''Takes the validation type assigned to the cell,
    cell value,  coordinates of the cell, errors of previous validation break list
    '''
    #Assigning each class to the validation type
    classmap = {
        'NotBlank': NotBlankValidator.NotBlankValidator,
        'Type': TypeValidator.TypeValidator,
        'Length': LengthValidator.LengthValidator,
        'Regex': RegexValidator.RegexValidator,
        'Email': EmailValidator.EmailValidator,
        'Choice': ChoiceValidator.ChoiceValidator,
        'Date': DateTimeValidator.DateTimeValidator,
        'ExcelDate': ExcelDateValidator.ExcelDateValidator,
        'Country': CountryValidator.CountryValidator,
        'Conditional': ConditionalValidator.ConditionalValidator
    }
    violations = []
    #name is the validation type name (NotBlank, Regex, Length ,.... etc)
    #data is the value assigned by the user to be validated ( 3 chars , regex pattern , ... etc)
    name = list(type.keys())[0]
    data =list(type.values())[0]
    validator = classmap[name](data)

    #conditional validator will take two arguments to evaluate

    if name != 'Conditional':
        result = validator.validate(value)
    else:
        result = validator.validate(value, value2)

    #If the cell value breakes the validation tule , append violations list
    if (result == False):
        violations.append(validator.getMessage())

    if len(violations) > 0:
        errors.append((coordinate, violations))

    #return result != False
    #result is the output of each validation for each cell
    if (result == False):
        return False
    else:
        return True

def setSettings(config):
    '''function takes the config yaml file and converts it to dictionary
    '''
    settings = {}

    #excludes are the columns that we won't validate

    excludes = []

    print ("Get validation config " + config)
    try:
        stream = open(config, 'r')
    except IOError as e:
        print (e)
        exit(1)
    config = yaml.safe_load(stream)

    #Make sure that the yaml file follows the rules
    if 'validators' in config and 'columns' in config.get('validators'):
        settings['validators'] = config.get('validators').get('columns')
    else:
        return False

    if 'default' in config.get('validators') :
        settings['defaultValidator'] = config.get('validators').get('default')[0]
    else:
        settings['defaultValidator'] = None

    if 'excludes' in config:
        for column in config.get('excludes'):
            excludes.append(column_index_from_string(column))
        settings['excludes'] = excludes
    else:
        settings['excludes'] = []

    if 'range' in config:
        settings['range'] = config.get('range')[0] + "1:" + config.get('range')[1]
    else:
        settings['range'] = None

    if 'header' in config:
        settings['header'] = config.get('header')
    else:
        settings['header'] = True

    return settings

def markErrors(errors, excelFile, sheetName, tmpDir, printErrors = False, noSizeLimit = False):
    ''' Function takes the error lists (coordinates,violations) , excel file , sheet name
    output directory
    '''
    progressBar = Bar('Processing', max = len(errors))

    if printErrors.lower() == 'true':
        print ("Log broken cells")
        for error in errors:
            progressBar.next()
            print (" Broken Excel cell: " + error[0] + " [ "+ ','.join(error[1]) + " ]")
        progressBar.finish()

        return

    #Checking size of the file
    isFileTooBig = os.path.getsize(excelFile) > 10485760

    if isFileTooBig is True and noSizeLimit.lower() != 'true':
        return -1

    #open Excel file
    newFile = os.path.join(tmpDir , "errors_" + time.strftime("%Y-%m-%d") + "_" + str(int(time.time())) + "_" + os.path.basename(excelFile))
    fileName,fileExtension = os.path.splitext(excelFile)

    if fileExtension == '.xlsm':
        wb = load_workbook(excelFile, keep_vba=True, data_only=True)
    else:
        wb = load_workbook(excelFile, data_only=True)

    creator = wb.properties.creator
    ws = wb.get_sheet_by_name(sheetName)

    #fill the error values with red pattern

    redFill = PatternFill(start_color='FFFF0000',
        end_color = 'FFFF0000',
        fill_type = 'solid')

    for error in errors:
        progressBar.next()

        print (" Broken Excel cell: " + error[0])
        cell = ws[error[0]]
        if printErrors:
            cell.value = ','.join(error[1])
        cell.fill = redFill

    progressBar.finish()

    #save error excel file
    wb.properties.creator = creator
    print ("[[Save file: " + newFile + "]]")

    try:
        wb.save(newFile)
    except Exception as e:
        print (e)
        exit(1)

    return newFile

def validate(settings, excelFile, sheetName, tmpDir, printErrors = False, noSizeLimit = False):
    '''the main function of valitations, takes settings dictionary (validations)
    and returns the validation result
    '''
    print ("Validate Excel Sheet " + sheetName)

    errors = []
    #open Excel file
    print ("Parse Excel file")
    wb = load_workbook(excelFile, keep_vba=True, data_only=True, read_only=True)
    #ws = wb.get_sheet_by_name(sheetName)
    ws = wb[sheetName]

    progressBar = Bar('Processing', max=ws.max_row)

    if 'range' in settings and settings['range'] != None:
        settings['range'] = settings['range'] + (str)(ws.max_row)
    # range now equals A1:D(150) for example

    #iterate excel sheet
    rowCounter = 0
    for row in ws.iter_rows(settings['range']):
        progressBar.next()
        columnCounter = 0
        rowCounter = rowCounter + 1
        #do not parse empty rows
        if isEmpty(row):
            continue
        for cell in row:
            columnCounter = columnCounter + 1
            try:
                value = cell.value
            except ValueError:
                #case when it is not possible to read value at all from any reason
                column = get_column_letter(columnCounter)
                coordinates = "%s%d" % (column, rowCounter)
                errors.append((coordinates, ValueError))

            #find header (first) row
            #if the code founded the first header "ID", then it is the header row
            if settings['header'] != True:

                if value == settings['header']:
                    #Replace the header with true incase it meets a value of "ID" downside
                    settings['header'] = True
                # skip el row
                break

            #skip excludes column
            if  hasattr(cell, 'column') and cell.column in settings['excludes']:
                continue

            column = get_column_letter(columnCounter)
            coordinates = "%s%d" % (column, rowCounter)

            ## column:A Coordinate:A2, for example

            if column in settings['validators']:
                for type in settings['validators'][column]:
                    name = list(type.keys())[0] # notblank, Regex, Length
                    if name != 'Conditional':
                        res = isValid(type, value, coordinates, errors)
                    else:
                        fieldB = list(type.values())[0]['fieldB']
                        value2 = ws[fieldB + str(rowCounter)].value
                        res = isValid(type, value, coordinates, errors, value2)
                    if not res:
                        break

            elif settings['defaultValidator'] != None:
                isValid(settings['defaultValidator'], value, coordinates, errors)

    progressBar.finish()

    print ("Found %d error(s)" % len(errors))
    if (len(errors) > 0):
        return markErrors(errors, excelFile, sheetName, tmpDir, printErrors, noSizeLimit)

    return True

def isEmpty(row):
    ''' function to get if the row is empty or not
    '''
    for cell in row:
        if cell.value:
            return False

    return True

if __name__ == '__main__':

    parser = argparse.ArgumentParser(description = 'Mark validation errors in Excel sheet.')
    parser.add_argument('config', metavar = 'config', help = 'Path to YAML config file')
    parser.add_argument('file', metavar = 'file', help = 'Path to excel sheet file')
    parser.add_argument('sheetName', metavar = 'sheetName', help = 'Excel Sheet Name')
    parser.add_argument('tmpDir', metavar = 'tmpDir', help = 'Temporary directory path')
    parser.add_argument('--errors', metavar = 'errors', help = 'Print errors messages without generating excel file with errors')
    parser.add_argument('--no-file-size-limit', metavar = 'size', help = 'Switch off file size limit. Use with care')
    args = parser.parse_args()

    settings = setSettings(args.config)

    if settings == False:
        sys.exit("Incorrect config file " + args.config)

    try:
        results = validate(settings, args.file, args.sheetName, args.tmpDir, args.errors, args.no_file_size_limit)
    except Exception as e:
        sys.exit("Error occured: " + str(e))

    # if result = True that means file is originaly true and all values are correct
    # if result != True and not equal None, get result file name
    # if results == -1 File is too large , Exit

    if results != True:
        if results and results != -1:
            sys.exit("Validation errors store in: [[" + results + "]]")
        elif results == -1:
            sys.exit("File is too big to generate annotated Excel file")

    sys.exit(0)
