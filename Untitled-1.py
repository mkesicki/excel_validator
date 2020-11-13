#def setSettings(config):
import yaml
import config
import openpyxl
from pprint import pprint
import argparse
import os.path
import sys
import time
import yaml
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from progress.bar import Bar
from validator import *
#import validators

settings = {}
excludes = []
config= "C:/Users/john.elmasry/excel_validator/example/example.yml"
print ("Get validation config " + config)
try:
    stream = open(config, 'r')
except IOError as e:
    print (e)
    exit(1)
config = yaml.safe_load(stream)

if 'validators' in config and 'columns' in config.get('validators'):
    settings['validators'] = config.get('validators').get('columns')
else:
    False

if 'default' in config.get('validators') :
    settings['defaultValidator'] = config.get('validators').get('default')[0]
else:
    settings['defaultValidator'] = None

if 'excludes' in config:
    for column in config.get('excludes'):
        excludes.append(openpyxl.utils.cell.column_index_from_string(column))
    settings['excludes'] = excludes
else:
    settings['excludes'] = []

if 'range' in config:
    settings['range'] = config.get('range')[0] + "1:" + config.get('range')[1]
else:
    settings['range'] = None

if 'header' in config:
    settings['header'] = config.get('header')
else:
    settings['header'] = True






classmap = {
        'NotBlank': NotBlankValidator.NotBlankValidator,
        'Type': TypeValidator.TypeValidator,
        'Length': LengthValidator.LengthValidator,
        'Regex': RegexValidator.RegexValidator,
        'Email': EmailValidator.EmailValidator,
        'Choice': ChoiceValidator.ChoiceValidator,
        'Date': DateTimeValidator.DateTimeValidator,
        'ExcelDate': ExcelDateValidator.ExcelDateValidator,
        'Country': CountryValidator.CountryValidator,
        'Conditional': ConditionalValidator.ConditionalValidator
}
violations = []
#type = settings.keys()[0]
#data = settings.values()[0]
type = list(settings.keys())[0]
data =list(settings.values())[0]
validator = classmap[type](data)

type.values()
classmap[type]
try : 
    6+"s"
except:
    for e in errors:
        print(e)

for type in settings['validators']["H"]:
    name = list(type.keys())[0]
    if name == 'Conditional':
        fieldB = list(type.values())[0]['fieldB']
        print(fieldB)

x